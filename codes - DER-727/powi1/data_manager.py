import pandas as pd
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series, BarChart
import os
import matplotlib.pyplot as plt
import shutil
from time import time, sleep
from gtts import gTTS
from playsound import playsound
from datetime import datetime

import getpass
username = getpass.getuser().lower()
# from datetime import datetime
# now = datetime.now()
# date = now.strftime('%m%d')


from openpyxl.styles import colors
from openpyxl.styles import Font, Color
# from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from openpyxl.styles import NamedStyle, Font, Border, Side

highlight = NamedStyle(name="highlight")
highlight.font = Font(name='Calibri', bold=False, size=11)
bd = Side(style='thin', color="000000")
highlight.alignment = Alignment(horizontal='center', vertical='center')
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

##########################################################################################################
def path_maker(file_path: str):
    """
        file_path: Enter the file path you want to create

        returns the string value of new path created
    """

    folder_list = file_path.split('/')

    new_path = ' '
    for i in folder_list:
        
        if new_path == ' ':
            path = f'{i}/'
        else: 
            path = new_path + f'{i}/'

        if not os.path.exists(path):
            os.mkdir(path)
            
        new_path = path
    return new_path

def remove_file(file_path: str):
    """
        file_path: Enter the file path you want to delete

        returns the string value of new path created
    """

    if os.path.exists(file_path): os.remove(file_path)

def move_file(source_path: str, destination_path: str):
    """
        source_path : path/file
        destination_path : path/file
    """
    source = f'{source_path}'
    destination = f'{destination_path}'
    remove_file(destination)
    print(f"{source} ==> ")
    print(f"{destination}")
    shutil.move(source, destination)
##########################################################################################################

##########################################################################################################
def sfx():
    import winsound as ws
    ws.PlaySound("dingding.wav", ws.SND_ASYNC)
    sleep(2) 

def create_sfx(message):
    language = 'en'
    myobj = gTTS(text=message, lang=language, slow=False)
    path_maker(f'{os.getcwd()}/sfx')
    if not os.path.exists(f"{os.getcwd()}/sfx/{message}.mp3"):
        myobj.save(f"{message}.mp3")
        move_file(f"{message}.mp3", f"{os.getcwd()}/sfx/{message}.mp3")
    playsound(f"{os.getcwd()}/sfx/{message}.mp3")

def tts(message):
    """
        function: translates text to speech based on user's input (message)
    """
    create_sfx(message)
    print(f"{message}")

def prompt(message):
    """
        function: translates text to speech based on user's input (message) and wait for user intervention to continue program
    """
    create_sfx(message)
    input(f"{message}")
###########################################################################################################

##########################################################################################################
def start_timer():
    start = datetime.now()
    return start

def end_timer(start):
    end = datetime.now()
    total_time = end-start
    temp = ""
    
    print(f"START TIME: {temp:>12}{start}")
    print(f"END TIME: {temp:>12}{end}")
    print(f"TOTAL TIME: {temp:>12}{total_time}")

def soak(soak_time):
    for seconds in range(soak_time, 0, -1):
        sleep(1)
        print(f"{seconds:5d}s", end="\r")
    print("       ", end="\r")
###########################################################################################################

##########################################################################################################
def headers(test_name):
    global start
    print()
    print("="*80)
    print(f"Test: {test_name}")    
    start = datetime.now()
    print("="*80)

def footers(waveform_counter=0):

    print("="*80)
    if waveform_counter > 0: print(f'{waveform_counter} waveforms captured.')
    print('test complete.')
    print()
    end = datetime.now()
    total_time = end-start
    temp = ""
    
    print(f"START TIME: {temp:>12}{start}")
    print(f"COMPLETION TIME: {temp:>7}{end}")
    print(f"TOTAL TESTING TIME: {temp:>16}{total_time}")
    print("="*80) 
##########################################################################################################    


##########################################################################################################
def convert_argv_to_int_list(a=[]):
    str_list = a.strip('[]').split(',')
    int_list = [int(x) for x in str_list]
    return int_list

def convert_argv_to_str_list(a=[]):
    str_list = a.strip('[]').split(',')
    return str_list
##########################################################################################################



##########################################################################################################
def get_anchor(col, row):
    """
    get anchor given a numerical col row  -> (col = 2, row = 4 -> 'B4')
    returns anchor (str)
    """
    anchor = f"{get_column_letter(col)}{row}"
    return anchor

def col_row_extractor(anchor):
    """
    extract col and row given an anchor (i.e. 'B4' -> col = 2, row = 4)

    anchor : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(anchor)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def get_df_col_len(df):
    df_row_len, df_col_len = df.shape
    return df_col_len

def get_df_row_len(df):
    df_row_len, df_col_len = df.shape
    return df_row_len

def change_anchor(anchor, row_increment=0, col_increment=0):
    """
    change anchor by incrementing either row or col then returns new equivalent anchor
    """
    col, row = col_row_extractor(anchor)
    row = row + row_increment
    col = col + col_increment
    new_anchor = get_anchor(col, row)
    return new_anchor

class default_constants():
    anchor = "B2"
    col_no = 0
    row_no = 0
    image_size_row = 34
    image_size_col = 17

def change_image_anchor(anchor, row_coords, col_coords):
    """
    change image anchor by specifying row and col coordinates
    """
    col, row = col_row_extractor(anchor)
    row = row + row_coords*default_constants.image_size_row
    col = col + col_coords*default_constants.image_size_col
    new_anchor = get_anchor(col, row)
    return new_anchor

##########################################################################################################

##########################################################################################################
### DF MANIPULATION CODES
##########################################################################################################
def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    sheet_list = wb.get_sheet_names()
    if sheet_name not in sheet_list: wb.create_sheet(sheet_name)
    try:
        default_sheet = wb.get_sheet_by_name('Sheet1')
        wb.remove_sheet(default_sheet)
    except: pass

    # print(sheet_list)
    if "Chart" not in sheet_list: wb.create_sheet("Chart")
    

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1

    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]
            try: wb[sheet_name][f'{get_column_letter(col)}{row}'].style = highlight
            except: wb[sheet_name][f'{get_column_letter(col)}{row}'].style = 'highlight'

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39 rows, 16 columns
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = folder_path + filename
    # file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1056
    img.height = 659.90551181

    sheet_list = wb.get_sheet_names()
    if sheet_name not in sheet_list: wb.create_sheet(sheet_name)
    try:
        default_sheet = wb.get_sheet_by_name('Sheet1')
        wb.remove_sheet(default_sheet)
    except: pass

    col, row = col_row_extractor(anchor)
    wb[sheet_name][f'{get_column_letter(col)}{row-1}'] = filename
    wb[sheet_name].add_image(img)

def create_header_list(df_header_list):
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    # print(df)
    return df

def export_to_excel(df, waveforms_folder, output_list, excel_name, sheet_name, anchor):
    df.loc[len(df)] = output_list
    # print(output_list)

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)

def export_df_to_excel(df, waveforms_folder, output_filename, sheet_name, anchor):

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{output_filename}"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)

    ## if top row is already the header
    df_header_list = list(df.columns.values)
    df_header = create_header_list(df_header_list)
    df_to_excel(wb, sheet_name, df_header, anchor)
    anchor = change_anchor(anchor, row_increment=1)

    df_to_excel(wb, sheet_name, df, anchor)
    
    wb.save(dst)

def export_screenshot_to_excel(output_filename, waveforms_folder, sheet_name, filename, anchor):
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{output_filename}"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    print(f"{filename} ==> {anchor}")
    wb = load_workbook(dst)
    image_to_excel(wb, sheet_name, filename=filename, folder_path=waveforms_folder, anchor=anchor)
    wb.save(dst)

    

def make_top_row_as_header_of_df(df):
    """
    @Summary
        make top row as header of dataframe
    @Description
    @Param
        df = dataframe
    @Returns
        df = dataframe
    @Example
    """
    df.columns = df.iloc[0] 
    df = df[1:]
    df.head()
    return df

def clean_dataframe(df):
    """
    @Summary
        Cleans raw dataframe
    @Description
        - removal of NaN rows
        - removal of NaN columns
        - removal of duplicate rows
        - setting of toprow as header of df
    @Param
        df = raw dataframe from read_excel
    @Returns
        df = clean dataframe
    @Example
    """
    df = df.dropna(axis=0, how='all') # dropping all NaN rows
    df = df.dropna(axis=1, how="all") # droping all NaN columns
    df = df.drop_duplicates() # dropping duplicate rows
    df = make_top_row_as_header_of_df(df)
    return df

def select_dataframe(df, start_anchor, end_anchor):
    """
    @Summary
        Select specific dataframe specified between the start_anchor and end_anchor region
    @Description
    @Param
        df = raw dataframe from read_excel
        start_anchor = upper left most corner of dataframe to be copied
        end_anchor = lower right most corner of dataframe to be copied
    @Returns
        df = selected dataframe region
    @Example
    """

    start_col, start_row = col_row_extractor(start_anchor)
    end_col, end_row = col_row_extractor(end_anchor)

    # df = df.iloc[:, start_col-1:end_col]
    df = df.iloc[start_row-2:end_row:, start_col-1:end_col]
    # df = df.loc[:, start_row:end_row]

    df = df.dropna(axis=0, how='all') # dropping all    NaN rows
    # df = df.dropna(axis=1, how="all") # droping all NaN columns
    df = df.drop_duplicates() # dropping duplicate rows
    # df = make_top_row_as_header_of_df(df)
    return df

def get_df_from_sheet(input_folder, input_filename, input_sheet_name):

    input_path = path_maker(f'{input_folder}')    
    input_file = input_folder + input_filename
    
    df = pd.read_excel(input_file, input_sheet_name)
    df = clean_dataframe(df)
    df = make_top_row_as_header_of_df(df)
    
    return df



def split_dataframe_by_unique_columns(df):
    """
    @Summary
        Separate multiple dataframe from a congregated dataframe by unique columns
    @Description
        Separate multiple dataframe from a congregated dataframe by unique columns
    @Param
        df = clean dataframe
    @Returns
        df_dict = dictionary of dataframes with index 0 being the left most part of the congregated dataframe
    """
    a = list(df.columns.values)    
    j = 0
    k = 0
    for i in a:
        if i == a[0] and k == 0:
            j += 1
            k += 1
        elif i != a[0]:
            j += 1
        elif i == a[0] and k != 0:
            break
    df_len = j
    df_count = int(len(a)/df_len)

    df_dict = {}
    df_counter = 0
    for i in range(df_count):
        start = df_counter*(df_len + 1)
        end = df_len*df_counter + df_len
        # df_list[f"{df_counter}"] = df.iloc[:, start:end]
        df_dict[i] = df.iloc[:, start:end]
        
        df_counter += 1

    return df_dict

def output_df_by_df_list_index_and_led_load(df, index, led):
    """
    @Assumptions
        Left most part of dataframe is the lower voltages
        This is a special one-use only function
    @Summary
        Output dataframe by input voltage and led load.
    @Description
        Output dataframe by input voltage and led load.
    @Param
        df = raw dataframe from read_excel
        vin = input voltage
        led = led voltage
    @Returns
        df = dataframe by specific input voltage and led condition
    """
    df = clean_dataframe(df)
    df_list = split_dataframe_by_unique_columns(df)
    df = df_list[str(index)]
    df = df.loc[df['led_load'] == led]
    
    return df
##########################################################################################################

##########################################################################################################
### CHARTING CODES
##########################################################################################################
def get_number_of_sheets(path, input_file):
    wb = openpyxl.load_workbook(path + input_file) 
    x = len(wb.sheetnames)
    return x

def get_sheet_names(path, input_file):
    wb = openpyxl.load_workbook(path + input_file) 
    sheet_name_list = wb.sheetnames
    return sheet_name_list

def create_scatter_chart(title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

    chart = ScatterChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale

    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.minorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.y_axis.minorUnit = y_minor_unit


    return chart

def create_bar_chart(title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

    chart = BarChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15
    chart.type = "col"

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale

    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.minorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.y_axis.minorUnit = y_minor_unit

    return chart

def reset_chartsheet(wb):
    chart_sheet = wb["Chart"]
    try:
        no_of_existing_charts = len(chart_sheet._charts)
        for i in range(no_of_existing_charts):
            del chart_sheet._charts[(i-1)]
    except: pass
    
    return chart_sheet

def save_chartsheet(chart_sheet, chart, chart_position):
    chart_sheet.add_chart(chart, chart_position)

def append_series(path, wb, ws_name, x_anchor, last_row_anchor, series_title, chart):
    ## used in der-727_chart_compiler.py
    
    df = excel_to_df(path, ws_name, x_anchor, last_row_anchor)
    ws = wb[ws_name]
    xcol, xrow = col_row_extractor(x_anchor)
    last_col, last_row = col_row_extractor(last_row_anchor)
    xvalues = Reference(ws, min_col=xcol, min_row=xrow, max_row=last_row)
    values = Reference(ws, min_col=last_col, min_row=xrow, max_row=last_row)
    series = Series(values, xvalues, title=series_title)
    series.marker=openpyxl.chart.marker.Marker('auto')
    series.graphicalProperties.line.noFill=False
    chart.series.append(series)


def append_df_to_series(path, wb, ws_name, anchor, x_axis_label, y_axis_label, df, series_title, chart):


    ws = wb[ws_name]

    ## computations
    xcol_anchor, xrow = col_row_extractor(anchor) # get reference anchor
    xrow = xrow + 1 # offset starting row so that header will not be included
    
    index_x = df.columns.get_loc(x_axis_label) # getting the index of from the dataframe where x-axis data series is located
    xcol = xcol_anchor + index_x # y-axis column

    index_y = df.columns.get_loc(y_axis_label) # getting the index of from the dataframe where y-axis data series is located
    ycol = xcol_anchor + index_y # y-axis column

    last_row = get_df_row_len(df) + 2 # last row for the chart

    # adding series on chart
    xvalues = Reference(ws, min_col=xcol, min_row=xrow, max_row=last_row)
    values = Reference(ws, min_col=ycol, min_row=xrow, max_row=last_row)
    series = Series(values, xvalues, title=series_title)
    series.marker=openpyxl.chart.marker.Marker('auto')
    series.graphicalProperties.line.noFill=False
    chart.series.append(series)
###########################################################################################################
