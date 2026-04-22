##################################################################################
from time import time, sleep

import math
import numpy as np
import pandas as pd
import openpyxl
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel
from powi.equipment import path_maker, remove_file, start_timer, end_timer
import os

import getpass
username = getpass.getuser().lower()
from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')
##################################################################################


class GENERAL_CONSTANTS():

    DEFAULT_WIDE_RANGE_VIN_LIST = [90,100,110,115,120,132,180,200,230,265,277,300]
    PPT_VIN_LIST = [90,115,230,265,277,300]

    SOAK_TIME_PER_LINE_QUICK_CHECK = 10
    SOAK_TIME_PER_LINE_DER = 300
    SOAK_TIME_PER_LINE_COMPARISON_CHECK = 180

    HEADER_LIST_CR_LOAD = ['CR (ohms)','Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)']
    # HEADER_LIST_CC_LOAD = ['Timestamp', 'CC (A)', 'Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)']
    HEADER_LIST_CC_LOAD = ['Vac (rms)',	'Freq (Hz)', 'Vin (rms)', 'Iin (mA)', 'Pin (W)', 'PF', '% THD', 'Vo (V)', 'Io (mA)', 'Po (W)', '%I Reg', 'Efficiency']
    HEADER_LIST_CC_LOAD_1 = ['Load%', 'Vac (rms)',	'Freq (Hz)', 'Vin (rms)', 'Iin (mA)', 'Pin (W)', 'PF', '% THD',
                             'Vo1 (V)', 'Io1 (mA)', 'Po1 (W)', '%I Reg1',
                             'Efficiency']
    HEADER_LIST_CC_LOAD_PDEL = ['Load%', 'Vac (rms)', 'Freq (Hz)', 'Vin (rms)', 'Iin (mA)', 'Pin (W)', 'PF', '% THD',
                             'Vo1 (V)', 'Io1 (mA)', 'Po1 (W)', '%V Reg1',
                             'Efficiency', "VoRipple"]
    HEADER_LIST_LED_LOAD = ['%Dim', 'Dim (V)', 'LED (V)'
                            'Vac (rms)', 'Freq (Hz)', 'Vin (rms)', 'Iin (mA)', 'Pin (W)', 'PF', '% THD',
                            'Vo1 (V)', 'Io1 (mA)', 'Po1 (W)', '%V Reg1',
                            'Efficiency']
    HEADER_LIST_CC_LOAD_2 = ['Load%', 'Vac (rms)',	'Freq (Hz)', 'Vin (rms)', 'Iin (mA)', 'Pin (W)', 'PF', '% THD',
                             'Vo1 (V)', 'Io1 (mA)', 'Po1 (W)', '%V Reg1',
                             'Vo2 (V)', 'Io2 (mA)', 'Po2 (W)', '%V Reg2',
                             'Overall Efficiency']
    HEADER_LIST_DIM = ['Timestamp', 'Dim (V)', 'CC (A)', 'Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)']
    HEADER_LIST_CC_LOAD_WIRELESS = ['Timestamp', 'CC (A)', 'Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)',
                                    'CE', 'CHS', 'TROUGH', 'PEAK',
                                    'VFOD_ADC_FILT', 'VFOD_ADC', 'VCOIL_ADC', 'DUTY_HOLDER', 'RP8',
                                    'PIN_ADC', 'PF_ADC', 'POUT_ADC', 'EFF_ADC', 'VIN_ADC', 'STATE',
                                    'POUT_COMMS_RECEIVED', 'FOD_STATUS']
    HEADER_LIST_CC_LOAD_VDS_STRESS = ['CC (A)', 'Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)', 'Vds_max']
    HEADER_LIST_CV_LOAD = ['CV (V)', 'Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)']
    # HEADER_LIST_LED_LOAD = ['LED (V)','Vin (rms)', 'Freq (Hz)', 'Vac (VAC)', 'Iin (A)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (A)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 

class GENERAL_FUNCTIONS():

    def CREATE_PATH(self, project, test):
        waveforms_folder = f'C:/Users/{username}/Desktop/{project}/{test}'
        path = path_maker(f'{waveforms_folder}')
        return waveforms_folder

    def PATH_MAKER(self, path_folder):
        path = path_maker(f'{path_folder}')
        return path

    def FILE_NAME_LIST(self, path_folder):
        """ returns a list of filenames present in the specified path folder
        """
        filename_list = os.listdir(path_folder)
        return filename_list
    
    def EXCEL_FILES_LIST(self, path_folder):
        """ returns only the list of excel files in the specified path folder
        """
        excel_files_list = []
        filename_list = self.FILE_NAME_LIST(path_folder)
        for file in filename_list:
            _, extension = os.path.splitext(file)
            if extension == '.xlsx': excel_files_list.append(file)
        return excel_files_list
    
    class EXCEL_FUNCTIONS():

        def GET_NUMBER_OF_SHEETS(self, path, input_file):
            wb = openpyxl.load_workbook(path + '/' + input_file) 
            x = len(wb.sheetnames)
            return x
        
        def GET_SHEET_NAMES_LIST(self, path, input_file):
            wb = openpyxl.load_workbook(path + '/' + input_file) 
            sheet_name_list = wb.sheetnames
            return sheet_name_list
        
        def _make_top_row_as_header_of_df(self, df):
            """
            @Summary
                make top row as header of dataframe
            @Description
            @Param
                df = dataframe
            @Returns
                df = dataframe
            @Example
            """
            df.columns = df.iloc[0] 
            df = df[1:]
            df.head()
            return df
        
        def CLEAN_DATAFRAME(self, df):
            """
                @Summary
                    Cleans raw dataframe
                @Description
                    - removal of NaN rows
                    - removal of NaN columns
                    - removal of duplicate rows
                    - setting of toprow as header of df
                @Param
                    df = raw dataframe from read_excel
                @Returns
                    df = clean dataframe
                @Example
            """
            df = df.dropna(axis=0, how='all') # dropping all NaN rows
            df = df.dropna(axis=1, how="all") # droping all NaN columns
            df = df.drop_duplicates() # dropping duplicate rows
            df = self._make_top_row_as_header_of_df(df)
            return df

        def CHANGE_ANCHOR(self, anchor, row_increment=0, col_increment=0):
            """
            change anchor by incrementing either row or col then returns new equivalent anchor
            """
            col, row = col_row_extractor(anchor)
            row = row + row_increment
            col = col + col_increment
            new_anchor = get_anchor(col, row)
            return new_anchor

        def EXPORT_DF_TO_EXCEL(self, df, input_folder, output_filename, sheet_name, anchor):

            src = f"{os.getcwd()}/blank.xlsx"
            dst = f"{input_folder}/{output_filename}"
            if not os.path.exists(dst): shutil.copyfile(src, dst)

            wb = load_workbook(dst)

            ## if top row is already the header
            df_header_list = list(df.columns.values)
            df_header = create_header_list(df_header_list)
            df_to_excel(wb, sheet_name, df_header, anchor)
            print(df_header)
            anchor = self.CHANGE_ANCHOR(anchor, row_increment=1)

            df_to_excel(wb, sheet_name, df, anchor)
            
            wb.save(dst)
        
        def RESET_CHARTSHEET(self, wb):
            chart_sheet = wb["Chart"]
            try:
                no_of_existing_charts = len(chart_sheet._charts)
                for i in range(no_of_existing_charts):
                    del chart_sheet._charts[(i-1)]
            except: pass
            
            return chart_sheet
        
        def CREATE_SCATTER_CHART(self, title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

            chart = ScatterChart()
            chart.title = title
            chart.style = style
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            chart.height = 10 # default is 7.5
            chart.width = 20 # default is 15

            chart.x_axis.scaling.min = x_min_scale
            chart.x_axis.scaling.max = x_max_scale
            chart.y_axis.scaling.min = y_min_scale
            chart.y_axis.scaling.max = y_max_scale

            chart.x_axis.majorUnit = x_major_unit
            chart.x_axis.minorUnit = x_minor_unit
            chart.y_axis.majorUnit = y_major_unit
            chart.y_axis.minorUnit = y_minor_unit


            return chart
    
        def APPEND_DF_TO_SERIES(self, path, wb, ws_name, anchor, x_axis_label, y_axis_label, df, series_title, chart):

            ws = wb[ws_name]

            ## computations
            xcol_anchor, xrow = col_row_extractor(anchor) # get reference anchor
            xrow = xrow + 1 # offset starting row so that header will not be included
            
            index_x = df.columns.get_loc(x_axis_label) # getting the index of from the dataframe where x-axis data series is located
            xcol = xcol_anchor + index_x # y-axis column

            index_y = df.columns.get_loc(y_axis_label) # getting the index of from the dataframe where y-axis data series is located
            ycol = xcol_anchor + index_y # y-axis column

            last_row = self._get_df_row_len(df) + 2 # last row for the chart

            # adding series on chart
            xvalues = Reference(ws, min_col=xcol, min_row=xrow, max_row=last_row)
            values = Reference(ws, min_col=ycol, min_row=xrow, max_row=last_row)
            series = Series(values, xvalues, title=series_title)
            series.marker=openpyxl.chart.marker.Marker('auto')
            series.graphicalProperties.line.noFill=False
            chart.series.append(series)

        def _get_df_col_len(self, df):
            df_row_len, df_col_len = df.shape
            return df_col_len

        def _get_df_row_len(self, df):
            df_row_len, df_col_len = df.shape
            return df_row_len
        
        def SAVE_CHARTSHEET(self, chart_sheet, chart, chart_position):
            chart_sheet.add_chart(chart, chart_position)

    def CREATE_DF_WITH_HEADER(self, header_list):
        """ creates a data frame with empty header_list specified by user
            returns a dataframe
        """
        return create_header_list(header_list)

    def PRINT_FINAL_DATA_DF(self, df):
        print(f"\n\nFinal Data: ")
        print(df)

    def ESTIMATED_TEST_TIME(self, estimated_time):
        print(f"Estimated Testing Time: {round(math.ceil(estimated_time/60), 2)} mins.")




    
