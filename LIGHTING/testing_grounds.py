import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import subprocess

def search_files():
    global folder_path
    if not folder_path or not os.path.exists(folder_path):
        folder_path = input("Enter folder path: ")
        with open("last_folder.txt", "w") as f:
            f.write(folder_path)
    else:
        print(f"Searching in folder: {folder_path}")

    keyword = input("Enter keyword: ").lower()
    print()

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"
    ws["A1"] = "File/Folder Name"
    ws["B1"] = "Path"

    files_found = False
    row = 2

    for root, dirs, files in os.walk(folder_path):
        dir_path = ""
        for dir in dirs:
            if keyword in dir.lower():
                dir_path = os.path.join(root, dir)
                ws.cell(row=row, column=1).value = dir
                ws.cell(row=row, column=2).value = dir_path
                ws.cell(row=row, column=2).hyperlink = f"file://{dir_path}"
                row += 1

        for file in files:
            if keyword in file.lower():
                file_path = os.path.join(root, file)
                ws.cell(row=row, column=1).value = file
                ws.cell(row=row, column=2).value = file_path
                ws.cell(row=row, column=2).hyperlink = f"file://{os.path.dirname(file_path)}"
                row += 1
                files_found = True

    if not files_found:
        print('No files found.')
    else:
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
            ws.cell(row=1, column=col).font = Font(bold=True)

        wb.save("C:/Users/ccayno/automation/codes/LIGHTING/file_searcher_outputs\search_results.xlsx")
        print("Search results saved to C:/Users/ccayno/automation/codes/LIGHTING/file_searcher_outputs/search_results.xlsx")

    print()
    search_again = input("Search again? (y/n) ").lower()
    if search_again == 'y':
        search_files()

if __name__ == '__main__':
    change_folder = input("Do you want to change the default folder? (y/n) ").lower()
    if change_folder == 'y':
        folder_path = None
    else:
        folder_path = None
        try:
            with open("last_folder.txt", "r") as f:
                folder_path = f.read()
        except FileNotFoundError:
            pass

    search_files()
