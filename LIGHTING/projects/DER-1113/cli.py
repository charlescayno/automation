#!/usr/bin/env python3
# ======================================================================================
# DER-1113 MASTER CLI TEST RUNNER - MULTI-UNIT EDITION
# ======================================================================================
# Project: DER-1113 Test Automation Suite
# Description: Centralized interactive menu, multi-unit test orchestrator, and
#              consolidated report generator for DER-1113 scripts.
# ======================================================================================

import sys
import os
import argparse
import subprocess
import time
import json
from datetime import datetime

# Path setup to ensure access to root libraries and venv
_current_dir = os.path.dirname(os.path.abspath(__file__))
_root_dir = os.path.abspath(os.path.join(_current_dir, "../.."))

_venv_python = os.path.abspath(os.path.join(_root_dir, "Scripts", "python.exe"))

if os.path.exists(os.path.join(_root_dir, "Lib", "site-packages")):
    sys.path.insert(0, os.path.join(_root_dir, "Lib", "site-packages"))
sys.path.insert(0, _root_dir)

# Terminal styling using colorama
try:
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError:
    class Fore:
        GREEN = RED = YELLOW = CYAN = MAGENTA = BLUE = WHITE = RESET = ""
    class Style:
        BRIGHT = RESET_ALL = ""

def title(msg):   print(f"{Fore.MAGENTA}{Style.BRIGHT}{msg}{Style.RESET_ALL}")
def info(msg):    print(f"{Fore.CYAN}{msg}{Style.RESET_ALL}")
def success(msg): print(f"{Fore.GREEN}{Style.BRIGHT}{msg}{Style.RESET_ALL}")
def warning(msg): print(f"{Fore.YELLOW}{msg}{Style.RESET_ALL}")
def error(msg):   print(f"{Fore.RED}{Style.BRIGHT}{msg}{Style.RESET_ALL}")

# ======================================================================================
# TEST SCRIPTS CATALOG
# ======================================================================================
CATALOG = [
    # Efficiency & Regulation
    {
        "id": "1",
        "key": "avg_efficiency",
        "name": "Average Efficiency (DOE6)",
        "file": "DER-1113 Average Efficiency.py",
        "category": "Efficiency & Regulation"
    },
    {
        "id": "2",
        "key": "line_load_reg",
        "name": "Line & Load Regulation",
        "file": "DER-1113 Line Load Regulation.py",
        "category": "Efficiency & Regulation"
    },
    {
        "id": "3",
        "key": "line_eff",
        "name": "Line Efficiency Sweep",
        "file": "DER-1113 Line Efficiency.py",
        "category": "Efficiency & Regulation"
    },
    {
        "id": "4",
        "key": "line_eff_simple",
        "name": "Line Efficiency (Simple)",
        "file": "DER-1113 Line Efficiency Simple.py",
        "category": "Efficiency & Regulation"
    },
    {
        "id": "5",
        "key": "line_eff_peak",
        "name": "Line Efficiency at Peak Power",
        "file": "DER-1113 Line Efficiency at Peak Power.py",
        "category": "Efficiency & Regulation"
    },
    {
        "id": "6",
        "key": "std_efficiency",
        "name": "Standard Efficiency Test",
        "file": "DER-1113_Efficiency.py",
        "category": "Efficiency & Regulation"
    },

    # Power & Protection
    {
        "id": "7",
        "key": "no_load",
        "name": "No Load Input Power",
        "file": "DER-1113 No Load Input Power.py",
        "category": "Power & Protection"
    },
    {
        "id": "8",
        "key": "brown_in_out",
        "name": "Brown In and Brown Out",
        "file": "DER-1113 Brown In and Brown Out.py",
        "category": "Power & Protection"
    },
    {
        "id": "9",
        "key": "ac_cycling",
        "name": "AC Mains Cycling",
        "file": "DER-1113 AC Cycling.py",
        "category": "Power & Protection"
    },
    {
        "id": "10",
        "key": "ac_transient",
        "name": "AC ON-OFF Transient",
        "file": "DER-1113 AC ON-OFF Transient.py",
        "category": "Power & Protection"
    },
    {
        "id": "11",
        "key": "peak_power",
        "name": "Peak Power Test",
        "file": "DER-1113 Peak Power.py",
        "category": "Power & Protection"
    },

    # Waveform Captures
    {
        "id": "12",
        "key": "pri_vds_ids_startup",
        "name": "Primary Vds/Ids Startup Waveform",
        "file": "DER-1113 Primary Vds Ids Startup Waveform.py",
        "category": "Waveforms & Scope Captures"
    },
    {
        "id": "13",
        "key": "pri_vds_ids_steady",
        "name": "Primary Vds/Ids Steady-State Waveform",
        "file": "DER-1113 Primary Vds Ids Steadystate Waveform.py",
        "category": "Waveforms & Scope Captures"
    },
    {
        "id": "14",
        "key": "diode_startup",
        "name": "Output Diode Startup Waveform",
        "file": "DER-1113 Output Diode Startup Waveforms.py",
        "category": "Waveforms & Scope Captures"
    },
    {
        "id": "15",
        "key": "diode_steady",
        "name": "Output Diode Steady-State Waveform",
        "file": "DER-1113 Output Diode Steady-state Waveform.py",
        "category": "Waveforms & Scope Captures"
    },
    {
        "id": "16",
        "key": "diode_peak_power",
        "name": "Output Diode Peak Power Waveform",
        "file": "DER-1113 Output Diode Peak Power.py",
        "category": "Waveforms & Scope Captures"
    },
    {
        "id": "17",
        "key": "output_ripple",
        "name": "Output Voltage Ripple Waveform",
        "file": "DER-1113_Output_Ripple.py",
        "category": "Waveforms & Scope Captures"
    }
]

# ======================================================================================
# HARDWARE DIAGNOSTIC CHECK
# ======================================================================================
def check_hardware():
    """Attempts connection to configured lab equipment and reports status."""
    title("\n========================================================")
    title("          PRE-FLIGHT HARDWARE DIAGNOSTIC CHECK          ")
    title("========================================================\n")

    try:
        from misc_codes.equipment_address import EQUIPMENT_ADDRESS
        info("Loaded instrument addresses from misc_codes.equipment_address:")
        print(f"  • Scope IP:               {getattr(EQUIPMENT_ADDRESS, 'SCOPE', 'N/A')}")
        print(f"  • AC Source GPIB:         {getattr(EQUIPMENT_ADDRESS, 'AC_SOURCE', 'N/A')}")
        print(f"  • E-Load GPIB:            {getattr(EQUIPMENT_ADDRESS, 'ELOAD', 'N/A')}")
        print(f"  • Power Meter Source GPIB:{getattr(EQUIPMENT_ADDRESS, 'POWER_METER_SOURCE', 'N/A')}")
        print(f"  • Power Meter Load 1 GPIB:{getattr(EQUIPMENT_ADDRESS, 'POWER_METER_LOAD_1', 'N/A')}")
        print(f"  • Signal Generator GPIB:  {getattr(EQUIPMENT_ADDRESS, 'SIG_GEN', 'N/A')}")
        print(f"  • DC Dimmer GPIB:         {getattr(EQUIPMENT_ADDRESS, 'DC_SOURCE_DIMMER', 'N/A')}")
        print()
    except Exception as e:
        error(f"Failed to load equipment_address module: {e}")

    info("Attempting PyVISA / Equipment Connections...")
    try:
        from misc_codes.equipment_settings import _connect, ACSource, PowerMeter, ElectronicLoad, Oscilloscope
        from misc_codes.equipment_address import EQUIPMENT_ADDRESS

        ac    = _connect("AC Source",          lambda: ACSource(EQUIPMENT_ADDRESS.AC_SOURCE))
        pms   = _connect("Power Meter Source", lambda: PowerMeter(EQUIPMENT_ADDRESS.POWER_METER_SOURCE))
        eload = _connect("Electronic Load",    lambda: ElectronicLoad(EQUIPMENT_ADDRESS.ELOAD))
        scope = _connect("Oscilloscope",       lambda: Oscilloscope(EQUIPMENT_ADDRESS.SCOPE))

        status_count = sum(1 for inst in [ac, pms, eload, scope] if inst is not None)
        if status_count == 4:
            success("\n[ALL CORE INSTRUMENTS CONNECTED SUCCESSFULLY]")
        elif status_count > 0:
            warning(f"\n[{status_count}/4 CORE INSTRUMENTS CONNECTED - SOME DISCONNECTED]")
        else:
            error("\n[NO INSTRUMENTS CONNECTED - CHECK GPIB/LAN CONNECTION]")
    except Exception as e:
        error(f"Diagnostic error: {e}")
    print()

# ======================================================================================
# SAFETY & HARDWARE SWAP ROUTINES
# ======================================================================================
def safe_hardware_discharge():
    """Calls DISCHARGE_OUTPUT safety routine between test runs."""
    try:
        from misc_codes.equipment_settings import EQUIPMENT_FUNCTIONS
        warning("Safety Action: Discharging E-Load Output...")
        EQUIPMENT_FUNCTIONS().DISCHARGE_OUTPUT(2)
        success("Output Discharged Safely.")
    except Exception as e:
        warning(f"Note: Could not run DISCHARGE_OUTPUT routine automatically: {e}")

def prompt_unit_swap(next_unit_id):
    """Prompts the operator to safely swap hardware for the next Unit under test."""
    safe_hardware_discharge()
    title("\n========================================================")
    title(f"   OPERATOR ACTION REQUIRED: SWAP DUT TO [{next_unit_id}]   ")
    title("========================================================")
    warning(f"1. Ensure AC Source is OFF.")
    warning(f"2. Disconnect previous unit from test fixture.")
    info(f"3. Connect new unit ({next_unit_id}) to AC Source, Power Meter, & E-Load.")
    title("========================================================")
    input(Fore.YELLOW + f"\nPress ENTER when [{next_unit_id}] is connected and ready to test..." + Style.RESET_ALL)
    success(f"Acknowledged. Commencing testing for [{next_unit_id}]...\n")

# ======================================================================================
# SCRIPT EXECUTION ENGINE WITH UNIT INJECTION
# ======================================================================================
def run_script(test_info, unit_id=None, auto_input=False):
    """Runs a single test script in a Python subprocess.
       If auto_input is False (default), standard input (stdin) remains connected to the live terminal,
       allowing the technician/operator to interactively enter user details, ambient temperature, test mode, etc.
    """
    script_path = os.path.join(_current_dir, test_info["file"])
    if not os.path.exists(script_path):
        error(f"Script file not found: {script_path}")
        return False, "File Not Found"

    unit_str = f" [{unit_id}]" if unit_id else ""
    title(f"\n========================================================")
    title(f" EXECUTING{unit_str}: {test_info['name']}")
    title(f" FILE:          {test_info['file']}")
    title(f" TIME:          {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    title(f"========================================================\n")

    start_time = time.time()
    
    # If auto_input is True, prepare automated input stream; otherwise allow live terminal input
    input_bytes = None
    if auto_input and unit_id:
        input_text = f"{unit_id}\n1\n1\n{unit_id}\n1\n1\n{unit_id}\n\n\n\n\n"
        input_bytes = input_text.encode('utf-8')

    raw_csv_filename = f"temp_raw_data_{test_info['id']}_{unit_id or 'default'}.csv"
    raw_csv_path = os.path.join(_current_dir, raw_csv_filename)
    if os.path.exists(raw_csv_path):
        try: os.remove(raw_csv_path)
        except Exception: pass

    env = os.environ.copy()
    if unit_id:
        env["DUT_UNIT_ID"] = unit_id
    env["RAW_DATA_EXPORT_PATH"] = raw_csv_path

    lib_sp = os.path.join(_root_dir, "Lib", "site-packages")
    current_pp = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = f"{lib_sp};{_root_dir};{_current_dir}" + (f";{current_pp}" if current_pp else "")

    py_exec = _venv_python if os.path.exists(_venv_python) else sys.executable

    raw_table_md = None
    try:
        result = subprocess.run(
            [py_exec, script_path],
            input=input_bytes,
            cwd=_current_dir,
            env=env,
            check=False
        )
        elapsed = time.time() - start_time

        # Process raw CSV data table if generated by the test script
        if os.path.exists(raw_csv_path):
            try:
                import pandas as pd
                df_raw = pd.read_csv(raw_csv_path)
                if not df_raw.empty:
                    headers = [str(c) for c in df_raw.columns]
                    md_rows = [
                        "| " + " | ".join(headers) + " |",
                        "| " + " | ".join(["---"] * len(headers)) + " |"
                    ]
                    for _, row in df_raw.iterrows():
                        row_vals = [str(val) for val in row.values]
                        md_rows.append("| " + " | ".join(row_vals) + " |")
                    raw_table_md = "\n".join(md_rows)
            except Exception:
                pass
            finally:
                try: os.remove(raw_csv_path)
                except Exception: pass

        if result.returncode == 0:
            success(f"\n[OK] [SUCCESS] {test_info['name']}{unit_str} completed in {elapsed:.1f}s")
            return True, f"{elapsed:.1f}s", raw_table_md
        else:
            code_hex = f"0x{result.returncode & 0xFFFFFFFF:08X}"
            reason = " (Instruments powered off or VISA driver timeout)" if result.returncode in [3221226505, -1073740791] else ""
            error(f"\n[FAIL] [FAILED] {test_info['name']}{unit_str} exited with code {code_hex}{reason} (took {elapsed:.1f}s)")
            return False, f"Error {code_hex}", raw_table_md
    except Exception as e:
        elapsed = time.time() - start_time
        error(f"\n[FAIL] [EXCEPTION] {test_info['name']}{unit_str} failed with error: {e}")
        return False, f"Exception: {e}", None

# ======================================================================================
# CONSOLIDATED REPORT PERSISTENT HISTORY MANAGEMENT
# ======================================================================================
HISTORY_FILE = os.path.join(_current_dir, "DER-1113_Consolidated_Data.json")

def load_consolidated_history():
    """Loads accumulated test history across all units from JSON storage."""
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            warning(f"Could not load history file ({e}). Starting fresh.")
    return {"units": {}}

def update_consolidated_history(new_batch_results):
    """Merges newly executed test results into persistent history without overwriting previous tests."""
    history = load_consolidated_history()
    if "units" not in history:
        history["units"] = {}

    for unit_id, test_list in new_batch_results.items():
        if unit_id not in history["units"]:
            history["units"][unit_id] = {}

        for test_res in test_list:
            test_name = test_res["test"]
            history["units"][unit_id][test_name] = {
                "file": test_res["file"],
                "status": test_res["status"],
                "duration": test_res["duration"],
                "raw_table": test_res.get("raw_table"),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=2)
    except Exception as e:
        error(f"Failed to update history JSON: {e}")

    return history

# ======================================================================================
# CONSOLIDATED MULTI-UNIT EXCEL REPORT GENERATOR
# ======================================================================================
def generate_multi_unit_consolidated_report(history_data):
    """Compiles/updates a consolidated summary Excel report across all accumulated tested units."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        warning(f"Could not load openpyxl ({e}). Skipping Excel consolidated report creation.")
        return

    report_filename = "DER-1113_Consolidated_Report.xlsx"
    report_path = os.path.join(_current_dir, report_filename)

    units_map = history_data.get("units", {})
    unit_ids = list(units_map.keys())
    if not unit_ids:
        return

    info(f"\nUpdating Consolidated Multi-Unit Excel Report: {report_filename}...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    # Styling definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    pass_font = Font(name="Calibri", size=11, bold=True, color="006100")
    fail_font = Font(name="Calibri", size=11, bold=True, color="9C0006")
    na_font = Font(name="Calibri", size=11, bold=False, color="595959")
    
    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Title Block
    ws.merge_cells(f"A1:{get_column_letter(4 + len(unit_ids))}1")
    title_cell = ws["A1"]
    title_cell.value = "DER-1113 MULTI-UNIT QUALIFICATION CONSOLIDATED REPORT"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Meta Info
    ws["A3"] = "Project Name:"
    ws["B3"] = "DER-1113"
    ws["A4"] = "Last Updated:"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Units Tested:"
    ws["B5"] = ", ".join(unit_ids)

    ws["A3"].font = bold_font
    ws["A4"].font = bold_font
    ws["A5"].font = bold_font

    # Table Header
    row_idx = 7
    headers = ["#", "Test Name", "Test Script File"] + unit_ids + ["Overall Status"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 25

    # Table Data Rows
    for t_idx, test in enumerate(CATALOG, 1):
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=t_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=test["name"]).font = bold_font
        ws.cell(row=row_idx, column=3, value=test["file"])

        unit_statuses = []
        col_offset = 4
        for u_id in unit_ids:
            u_data = units_map.get(u_id, {})
            test_res = u_data.get(test["name"])
            st = test_res["status"] if test_res else "N/A"
            unit_statuses.append(st)

            cell = ws.cell(row=row_idx, column=col_offset, value=st)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if st == "PASS":
                cell.font = pass_font; cell.fill = pass_fill
            elif st == "FAIL":
                cell.font = fail_font; cell.fill = fail_fill
            else:
                cell.font = na_font
            col_offset += 1

        if all(s == "PASS" for s in unit_statuses if s != "N/A") and any(s == "PASS" for s in unit_statuses):
            overall_st = "PASS"
        elif any(s == "FAIL" for s in unit_statuses):
            overall_st = "FAIL"
        else:
            overall_st = "N/A"

        ov_cell = ws.cell(row=row_idx, column=col_offset, value=overall_st)
        ov_cell.alignment = Alignment(horizontal="center")
        ov_cell.border = thin_border
        if overall_st == "PASS":
            ov_cell.font = pass_font; ov_cell.fill = pass_fill
        elif overall_st == "FAIL":
            ov_cell.font = fail_font; ov_cell.fill = fail_fill
        else:
            ov_cell.font = na_font

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    try:
        wb.save(report_path)
        success(f"[OK] [CONSOLIDATED EXCEL REPORT UPDATED] {report_path}")
    except Exception as e:
        error(f"Failed to save consolidated Excel report: {e}")

def generate_multi_unit_markdown_report(history_data):
    """Compiles/updates a consolidated summary Markdown report across all accumulated tested units."""
    report_filename = "DER-1113_Consolidated_Report.md"
    report_path = os.path.join(_current_dir, report_filename)

    units_map = history_data.get("units", {})
    unit_ids = list(units_map.keys())
    if not unit_ids:
        return

    info(f"Updating Consolidated Multi-Unit Markdown Report: {report_filename}...")

    lines = []
    lines.append("# DER-1113 Multi-Unit Qualification Consolidated Report\n")
    lines.append(f"**Project:** DER-1113 Test Automation  ")
    lines.append(f"**Report Last Updated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ")
    lines.append(f"**Units Tested:** {', '.join(unit_ids)}  \n")

    lines.append("## Executive Summary Matrix\n")

    # Table Header
    header_cols = ["#", "Test Name", "Test File"] + unit_ids + ["Overall Status"]
    lines.append("| " + " | ".join(header_cols) + " |")
    lines.append("| " + " | ".join(["---"] * len(header_cols)) + " |")

    # Table Rows
    for idx, test in enumerate(CATALOG, 1):
        row_cells = [str(idx), f"**{test['name']}**", f"`{test['file']}`"]
        unit_statuses = []

        for u_id in unit_ids:
            u_data = units_map.get(u_id, {})
            test_res = u_data.get(test["name"])
            st = test_res["status"] if test_res else "N/A"
            unit_statuses.append(st)

            if st == "PASS":
                row_cells.append("🟢 PASS")
            elif st == "FAIL":
                row_cells.append("🔴 FAIL")
            else:
                row_cells.append("⚪ N/A")

        if any(s == "FAIL" for s in unit_statuses):
            overall_icon = "🔴 FAIL"
        elif any(s == "PASS" for s in unit_statuses) and all(s in ["PASS", "N/A"] for s in unit_statuses):
            overall_icon = "🟢 PASS"
        else:
            overall_icon = "⚪ N/A"

        row_cells.append(overall_icon)
        lines.append("| " + " | ".join(row_cells) + " |")

    lines.append("\n---\n")

    # Detailed Results & Raw Measurement Data Per Unit
    lines.append("## Detailed Test Results & Raw Measurements Per Unit\n")
    for u_id, u_tests in units_map.items():
        lines.append(f"### Unit: {u_id}\n")
        if not u_tests:
            lines.append("*No completed tests recorded for this unit.*\n")
            continue

        for idx, (t_name, res) in enumerate(u_tests.items(), 1):
            st_icon = "🟢 PASS" if res["status"] == "PASS" else "🔴 FAIL"
            lines.append(f"#### {idx}. {t_name}")
            lines.append(f"- **Test File:** `{res['file']}`")
            lines.append(f"- **Status:** {st_icon}")
            lines.append(f"- **Duration:** {res['duration']}")
            lines.append(f"- **Last Updated:** {res.get('updated_at', 'N/A')}\n")
            if res.get("raw_table"):
                lines.append("##### Raw Measured Test Data:\n")
                lines.append(res["raw_table"])
                lines.append("\n")
            else:
                lines.append("*No raw tabular data captured for this test run.*\n")

    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        success(f"[OK] [CONSOLIDATED MD REPORT UPDATED] {report_path}")
    except Exception as e:
        error(f"Failed to save consolidated Markdown report: {e}")

# ======================================================================================
# MULTI-UNIT BATCH RUNNER
# ======================================================================================
def run_multi_unit_batch(units_list, tests_to_run, auto_input=False):
    """Runs a batch test suite sequentially across multiple units under test."""
    title("\n========================================================")
    title(f"     STARTING MULTI-UNIT BATCH SUITE RUN               ")
    title(f"     UNITS TO TEST ({len(units_list)}): {', '.join(units_list)}")
    title(f"     TESTS PER UNIT ({len(tests_to_run)}):")
    for t in tests_to_run:
        info(f"       • {t['name']}")
    title("========================================================\n")

    all_unit_results = {}
    suite_start = time.time()

    for u_idx, unit_id in enumerate(units_list, 1):
        if u_idx > 1:
            prompt_unit_swap(unit_id)

        title(f"\n>>> COMMENCING TESTING FOR UNIT [{u_idx}/{len(units_list)}]: {unit_id} <<<")
        unit_results = []

        for t_idx, test in enumerate(tests_to_run, 1):
            info(f"\n[{unit_id}] Running Test [{t_idx}/{len(tests_to_run)}]: {test['name']}")
            
            status, duration, raw_table = run_script(test, unit_id=unit_id, auto_input=auto_input)
            unit_results.append({
                "test": test["name"],
                "file": test["file"],
                "status": "PASS" if status else "FAIL",
                "duration": duration,
                "raw_table": raw_table
            })

            if t_idx < len(tests_to_run):
                safe_hardware_discharge()
                time.sleep(2)

        all_unit_results[unit_id] = unit_results

    total_elapsed = time.time() - suite_start

    # Print Summary Table
    title("\n========================================================")
    title("          MULTI-UNIT EXECUTION SUMMARY TABLE            ")
    title("========================================================")
    header_str = f"{'#':<3} | {'Test Name':<35}"
    for u in units_list:
        header_str += f" | {u:<8}"
    header_str += " | Overall"
    print(header_str)
    print("-" * len(header_str))

    total_passed = 0
    total_tests_count = len(tests_to_run) * len(units_list)

    for idx, test in enumerate(tests_to_run, 1):
        row_str = f"{idx:<3} | {test['name']:<35}"
        test_passed_across_units = True

        for u_id in units_list:
            res = next((r for r in all_unit_results[u_id] if r["test"] == test["name"]), None)
            st = res["status"] if res else "N/A"
            if st == "PASS":
                total_passed += 1
                row_str += f" | {Fore.GREEN}PASS{Style.RESET_ALL}    "
            else:
                test_passed_across_units = False
                row_str += f" | {Fore.RED}FAIL{Style.RESET_ALL}    "

        ov_str = f"{Fore.GREEN}PASS{Style.RESET_ALL}" if test_passed_across_units else f"{Fore.RED}FAIL{Style.RESET_ALL}"
        row_str += f" | {ov_str}"
        print(row_str)

    print("-" * len(header_str))
    print(f"Total: {total_passed}/{total_tests_count} Individual Unit Tests Passed in {total_elapsed/60:.1f} minutes.\n")

    # Update persistent merged history and generate reports (preserves & updates existing data)
    accumulated_history = update_consolidated_history(all_unit_results)
    generate_multi_unit_consolidated_report(accumulated_history)
    generate_multi_unit_markdown_report(accumulated_history)

# ======================================================================================
# INTERACTIVE MENU
# ======================================================================================
def display_interactive_menu():
    """Main interactive terminal interface."""
    while True:
        title("\n========================================================")
        title("         DER-1113 MASTER CLI - MULTI-UNIT EDITION       ")
        title("========================================================")

        current_category = None
        for item in CATALOG:
            if item["category"] != current_category:
                current_category = item["category"]
                info(f"\n--- {current_category.upper()} ---")
            print(f"  [{item['id']:>2}] {item['name']:<38} ({item['file']})")

        info("\n--- SUITES & MULTI-UNIT UTILITIES ---")
        print(f"  [ M] Multi-Unit Test Suite Execution")
        print(f"  [ A] Run ALL Tests for Single Unit")
        print(f"  [ H] Hardware Diagnostic Check")
        print(f"  [ Q] Quit")

        choice = input(Fore.YELLOW + "\nSelect an option (1-17, M, A, H, Q): " + Style.RESET_ALL).strip().upper()

        if choice == 'Q':
            info("Exiting DER-1113 Master CLI. Goodbye!")
            break
        elif choice == 'H':
            check_hardware()
        elif choice == 'M':
            title("\n================ MULTI-UNIT SETUP ================")
            units_in = input("Enter Unit IDs separated by commas (e.g. Unit 1, Unit 2, Unit 3): ").strip()
            if not units_in:
                error("No units specified. Aborting.")
                continue
            units = [u.strip() for u in units_in.split(",") if u.strip()]

            title("\nSelect tests to run across all units:")
            print("  [A] Full Qualification Suite (All 17 Tests)")
            print("  [C] Custom Test Numbers (e.g. 1, 2, 7)")
            sub_choice = input("Enter choice (A or C): ").strip().upper()

            if sub_choice == 'A':
                run_multi_unit_batch(units, CATALOG)
            elif sub_choice == 'C':
                t_nums = input("Enter test numbers separated by commas (e.g., 1, 2, 7): ").strip()
                t_indices = [t.strip() for t in t_nums.split(",") if t.strip()]
                selected = [item for item in CATALOG if item["id"] in t_indices]
                if selected:
                    run_multi_unit_batch(units, selected)
                else:
                    error("Invalid test selections.")
        elif choice == 'A':
            unit_id = input("Enter Unit ID (default: Unit 1): ").strip() or "Unit 1"
            confirm = input(Fore.RED + f"Run ALL 17 tests for [{unit_id}]? (y/N): " + Style.RESET_ALL).strip().lower()
            if confirm == 'y':
                run_multi_unit_batch([unit_id], CATALOG)
        else:
            choices = [c.strip() for c in choice.split(",") if c.strip()]
            selected_tests = []
            for c in choices:
                match = next((item for item in CATALOG if item["id"] == c), None)
                if match:
                    selected_tests.append(match)
                else:
                    error(f"Invalid option: '{c}'")

            if selected_tests:
                unit_id = input("Enter Unit ID (default: Unit 1): ").strip() or "Unit 1"
                run_multi_unit_batch([unit_id], selected_tests)

# ======================================================================================
# MAIN CLI ARGUMENT PARSER
# ======================================================================================
def main():
    parser = argparse.ArgumentParser(
        description="DER-1113 Master CLI Test Automation Orchestrator - Multi-Unit Edition",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python cli.py                                    # Launch interactive menu
  python cli.py --list                             # List all available test scripts
  python cli.py --check-hardware                   # Run equipment connection check
  python cli.py --units Unit1,Unit2 --run 1,2,7    # Run tests 1, 2, 7 across Unit 1 and Unit 2
  python cli.py --units Unit1,Unit2,Unit3 --run all# Run full suite across 3 units
"""
    )

    parser.add_argument("--list", "-l", action="store_true", help="List all available DER-1113 test scripts")
    parser.add_argument("--check-hardware", "-ch", action="store_true", help="Run hardware diagnostic check")
    parser.add_argument("--units", "-u", type=str, help="Comma-separated list of Unit IDs (e.g., Unit1,Unit2,Unit3)")
    parser.add_argument("--run", "-r", type=str, help="Run test(s) by ID, key, or 'all' (comma-separated)")

    args = parser.parse_args()

    if args.list:
        title("\nAVAILABLE DER-1113 TEST SCRIPTS:")
        for item in CATALOG:
            print(f"  ID: {item['id']:>2} | Key: {item['key']:<20} | {item['name']} ({item['file']})")
        print()
        return

    if args.check_hardware:
        check_hardware()
        return

    if args.run:
        units_list = [u.strip() for u in args.units.split(",") if u.strip()] if args.units else ["Unit 1"]
        
        if args.run.lower() == "all":
            selected_tests = CATALOG
        else:
            keys = [k.strip().lower() for k in args.run.split(",") if k.strip()]
            selected_tests = []
            for k in keys:
                match = next((item for item in CATALOG if item["id"] == k or item["key"].lower() == k), None)
                if match:
                    selected_tests.append(match)
                else:
                    error(f"Test key/ID '{k}' not found in catalog. Use --list to view valid options.")

        if selected_tests:
            run_multi_unit_batch(units_list, selected_tests, auto_input=True)
        return

    display_interactive_menu()

if __name__ == "__main__":
    main()
