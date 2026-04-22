from functools import wraps
from math import isnan
from time import sleep, time
from abc import ABC, abstractmethod
import atexit
import os
import sys
import shutil


try:
    from pyfirmata import Arduino, util
    import pyfirmata
    import pyvisa
    import numpy as np
    from gtts import gTTS
    import sys
    from playsound import playsound
    import pandas as pd
    from openpyxl import Workbook
    import openpyxl
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    from openpyxl.utils import get_column_letter
    import os
    import matplotlib.pyplot as plt


except:
    import pip
    pip.main(['install','pyqt5'])
    pip.main(['install','pyinstaller'])
    pip.main(['install','pyautogui'])
    pip.main(['install','pyfirmata'])
    pip.main(['install', 'pyvisa'])
    pip.main(['install','pandas'])
    pip.main(['install','openpyxl'])
    pip.main(['install','opencv-python'])
    pip.main(['install','matplotlib'])
    pip.main(['install','numpy'])
    pip.main(['install', 'gTTS'])
    pip.main(['install', 'playsound'])

    from pyfirmata import Arduino, util
    import pyfirmata
    import pyvisa
    import numpy as np
    pass


## FOR FORMATTING WHEN TRANSFERRING DF TO EXCEL
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Font, Border, Side
highlight = NamedStyle(name="highlight")
highlight.font = Font(name='Calibri', bold=False, size=11)
bd = Side(style='thin', color="000000")
highlight.alignment = Alignment(horizontal='center', vertical='center')
# highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)


GPIB_NO = 0
PORT = 0

type_to_address = {
    'gpib': f'GPIB{GPIB_NO}',
    'lan': 'TCPIP',
    'lan_lecroy': 'TCPIP',
    'usb': 'USB',
    'serial': 'ASRL'
    }

def reject_nan(func):
    @wraps(func)
    def wrapped(*args, **kwargs):
        for _ in range(10):
            response = func(*args, **kwargs)
            if not isnan(response):
                return response
            sleep(0.2)
        return None
    return wrapped

class Equipment(ABC):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        rm = pyvisa.ResourceManager()
        # print(rm.list_resources())

        if comm_type == 'serial':
            self.address = f'{type_to_address[comm_type]}{address}'
            # print(self.address)
        elif comm_type == 'lan_lecroy':
            self.address = f'{type_to_address[comm_type]}::{address}::INSTR'
        else:
            self.address = f'{type_to_address[comm_type]}::{address}'
        self.timeout = timeout
        try:
            self.device = rm.open_resource(self.address)
            # print(self.device.query('*IDN?'))
            self.device.timeout = self.timeout
            if not comm_type == 'lan_lecroy':
                atexit.register(self.cleanup)
        except Exception as e:
            raise pyvisa.VisaIOError(e)
            
        self._id = 0

    def __repr__(self):
        return (f'{self.__class__.__name__}'
                 '(address={self.address}, timeout={self.timeout})')

    @abstractmethod
    def cleanup(self):
        pass

    @property
    def id(self):
        self._id = self.write('*IDN?')
        return self._id

    def close(self):
        self.device.close()

    # def write(self, command):
    #     response = None
    #     try:
    #         if "?" in command:
    #             response = self.device.query(command).strip()
    #         else:
    #             self.device.write(command)
    #     except Exception as e:
    #         raise pyvisa.VisaIOError(e)

    #     return response
    
    def write(self, command):
        response = None
        reply_available = False # CDO
        try:
            if "BDMM" in command: # CDO
                reply_available = True
                command = command.split(':')[1]

            if "?" in command:
                response = self.device.query(command).strip()
            else:
                self.device.write(command)

            if reply_available: # CDO
                self.device.read()
                reply_available = False

        except Exception as e:
            raise pyvisa.VisaIOError(e)

        return response


class Oscilloscope(Equipment):
    def __init__(self, address, comm_type='lan', timeout=1E6):
        super().__init__(address, comm_type, timeout)

        self.write('SYST:DISP:UPD 1')


    def _extract_png(self, raw):
        offset = int(raw[1])-ord('0')
        return raw[offset+2:]

    def get_screenshot(self, filename="default.png", path=os.path.dirname(os.path.realpath(__file__))):
        self.write("HCOP:DEST \'MMEM\'")
        self.write("HCOP:DEV:LANG PNG")
        self.write("HCOP:DEV:INV ON")
        self.write("MMEM:NAME \'C:\\HCOPY.png\'")
        self.write("HCOP:IMMediate; *OPC?")

        self.device.write("MMEM:DATA? \'C:\\HCOPY.png\'")
        raw_image = self.device.read_raw()
        image = self._extract_png(raw_image)

        with open(path + "\\" + filename, "wb") as f:
            f.write(image)
            
    def _extract_channel_data(self, channel, raw):
        scale = self.channel[channel].scale
        position = self.channel[channel].position
        offset = self.channel[channel].offset
        factor = scale * 10 / (253 * 256)
        border = self.device.read("FORM:BORD?").strip()

        num_digits = raw[1] - 48
        num_values = int(raw[2 : 2 + num_digits])
        data = raw[2 + num_digits : 2 + num_digits + num_values]
        if border == "LSBF":
            adc = np.array(
                [
                    int.from_bytes(data[i : i + 2], "little")
                    for i in range(0, len(data), 2)
                ],
                dtype=np.int16,
            )
        else:
            adc = np.array(
                [
                    int.from_bytes(data[i : i + 2], "big")
                    for i in range(0, len(data), 2)
                ],
                dtype=np.int16,
            )
        channel_data = adc * factor - scale * position + offset
        return channel_data

    
    def save_channel_data(self, channel=1):
        self.write("EXPort:WAVeform:MULTIchannel ON")
        self.write("EXP:WAV:INCX OFF")
        for i in range(1, 5):
            if i == channel:
                self.write(f"CHANnel{i}:EXPortstate ON")
            else:
                self.write(f"CHANnel{i}:EXPortstate OFF")
        self.write("FORM:DATA INT,16")
        self.write("FORM ASC")
        a=self.write(f"CHAN{channel}:WAV1:DATA?")
        return a
        # input("tama naalskjdflakjsdfljkasdf")
        # raw_data = self.device.read()
        # print(raw_data)
        # return self._extract_channel_data(channel, raw_data)

    def probe_state(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:STATe?')

    def probe_type(self, channel=1):   
        return self.write(f'PROBe{channel}:SETup:TYPE?')

    def probe_name(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:NAME?')

    def probe_bandwidth(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:BANDwidth?')

    def probe_attenuation(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:ATTenuation[:AUTO]?')


    def auto_zero(self, channel):
        self.write(f"PROBe{channel}:SETup:OFFSet:AZERo")


    def get_measure(self, channel=1):
        channel_state = self.write(f'MEAS{channel}:ENAB?')
        if channel_state == '0':
            return None, None
        labels = []
        values = []
        self.write(f'MEAS{channel}:ARN ON')
        for item in self.write(f'MEAS{channel}:ARES?').split(','):
            label, value = item.split(':')
            labels.append(label.strip())
            values.append(float(value.strip()))        

        return labels, values
    
    def get_measure_dict(self, channel=1):
        channel_state = self.write(f'MEAS{channel}:ENAB?')
        if channel_state == '0':
            return None
        result = {}
        self.write(f'MEAS{channel}:ARN ON')
        for item in self.write(f'MEAS{channel}:ARES?').split(','):
            label, value = item.split(':')
            result[label.strip()] = float(value.strip())
        return result

    def get_measure_all(self):
        result = []
        for i in range(1, 9):
            labels, values = self.get_measure(i)
            result.append({
                "channel": i,
                "labels": labels,
                "values": values
            })
        return result

    def get_vertical(self, channel=1):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        result = {
            "channel": str(channel),
            "scale": float(self.write(f'CHAN{channel}:SCAL?')),
            "position": float(self.write(f'CHAN{channel}:POS?')),
            "offset": float(self.write(f'CHAN{channel}:OFFS?')),
            "coupling": self.write(f'CHAN{channel}:COUP?'),
            "bandwidth": self.write(f'CHAN{channel}:BAND?')
        }
        return result

    def get_horizontal(self):
        result = {
            "scale": float(self.write('TIM:SCAL?')),
            "position": float(self.write('TIM:HOR:POS?')),
            "resolution": float(self.write('ACQ:RES?')),
            "sample rate": float(self.write('ACQ:SRAT?'))
        }
        return result
      

    def get_cursor(self, cursor=1):
        cursor_state = self.write(f'CURS{cursor}:STAT?')
        if cursor_state == '0':
            return None

        result = {
            "x1 position": float(self.write(f'CURS{cursor}:X1P?')),
            "x2 position": float(self.write(f'CURS{cursor}:X2P?')),
            "y1 position": float(self.write(f'CURS{cursor}:Y1P?')),
            "y2 position": float(self.write(f'CURS{cursor}:Y2P?')),
            "delta x": float(self.write(f'CURS{cursor}:XDEL?')),
            "delta y": float(self.write(f'CURS{cursor}:YDEL?')),
            "source": self.write(f'CURS{cursor}:SOUR?')
        }
        return result

    def run(self):
        self.write('RUN')
        self.write('DISP:TRIG:LIN OFF')

    def run_single(self):
        self.trigger_mode(mode='NORM')
        # self.trigger_mode(mode='AUTO')
        self.write('RUNS')
        self.write('DISP:TRIG:LIN OFF')

    def stop(self):
        self.write('STOP')
        self.write('DISP:TRIG:LIN OFF')

    # CMC Code Update #################

    def record_length(self, record_length):
        """
            record_length : 1000 to 1 000 000 000
        """
        self.write(f'ACQ:POIN {record_length}')
        # print('Record Length: '+ str(record_length) + ' Sa') 

    def resolution(self, resolution):
        self.write(f'ACQ:RES {resolution}')
        print('Resolution: '+ str(resolution) + ' s') # 1E-15 to 0.5

    def time_position(self, time_position):
        self.write(f'TIM:REF {time_position}')
        self.write('TIM:HOR:POS 0')
    
    def time_scale(self, time_scale):
        """
        Parameters:
        <TimeScale> Range: 25E-12 to 50
                    Increment: 1E-12
                    *RST: 10E-9
                    Default unit: s/div
        """
        self.write(f'TIM:SCAL {time_scale}')

    def position_scale(self, time_position, time_scale):
        self.time_position(time_position)
        self.time_scale(time_scale)

    def remove_zoom(self):
        self.write("LAYout:ZOOM:REM 'Diagram1', 'Zoom1'")
        self.write("LAYout:ZOOM:REM 'Diagram1', 'Zoom2'")
        # self.remove_zoom_gate()

    def remove_zoom_gate(self):
        self.write("MEASurement1:GATE OFF")
        self.write(f"MEASurement1:GATE:ZCOupling OFF")

    def add_zoom_gate(self):
        self.write("MEASurement1:GATE ON")
        self.write(f"MEASurement1:GATE:ZCOupling ON")


    def add_zoom(self, rel_pos=50, rel_scale=10):
        self.remove_zoom()
        for i in range(5):
            self.write(f"LAYout:ZOOM:ADD 'Diagram{i}', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
            time_span = float(self.get_horizontal()['scale'])*10
            # print(f"Zoom scale: {((rel_scale/100)*time_span)/10} s/div.")
            self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram{i}', 'Zoom1', {rel_scale}")
            self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram{i}', 'Zoom1', {rel_pos}")
            self.write(f"LAYout:ZOOM:VERT:REL:SPAN 'Diagram{i}', 'Zoom1', MAX")
            # self.write("MEASurement1:GATE ON")
            # self.write(f"MEASurement1:GATE:ZCOupling ON")
    
    def add_zoom_with_gate(self, rel_pos=50, rel_scale=10):
        self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
        time_span = float(self.get_horizontal()['scale'])*10
        # print(f"Zoom scale: {((rel_scale/100)*time_span)/10} s/div.")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom1', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom1', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom1', 100")
        self.add_zoom_gate()
        
    
    def add_zoom_1(self, rel_pos=50, zoom_scale=6):
        # self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
        time_span = float(self.get_horizontal()['scale'])*10
        print(f"Zoom scale: {zoom_scale} s/div.")
        rel_scale = (zoom_scale/time_span)*100*10
        print(f"Rel Scale: {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom1', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom1', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom1', 100")
        # self.write("MEASurement1:GATE ON")
        # self.write(f"MEASurement1:GATE:ZCOupling ON")
    
    def add_zoom_2(self, rel_pos=50, zoom_scale=6):
        # self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom2'")
        time_span = float(self.get_horizontal()['scale'])*10
        print(f"Zoom scale: {zoom_scale} s/div.")
        rel_scale = (zoom_scale/time_span)*100*10
        print(f"Rel Scale: {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom2', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom2', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom2', 100")
        # self.write("MEASurement1:GATE ON")
        # self.write(f"MEASurement1:GATE:ZCOupling ON")

    def edge_trigger(self, trigger_channel, trigger_level, trigger_edge):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4

        self.trigger_mode(mode='NORM')
        
        # # setting trigger to absolute without hysteresis
        # self.write(f"TRIGger:LEVel{trigger_channel}:NOISe MANual")
        # self.write(f'TRIGger:LEVel{trigger_channel}:NOISe:ABSolute 0')

        self.write(f'TRIG1:SOUR {channel}')
        self.write('TRIG1:TYPE EDGE')
        self.write(f'TRIG1:LEV{trigger_channel} {trigger_level}') # range: -10 to 10, increment: 1E-3
        self.write(f'TRIG1:EDGE:SLOP {trigger_edge}') # POS | NEG | EITH
    
    def width_trigger(self, trigger_channel, width_polarity='POS', width_range='LONG', width=100E-3, delta=0):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4
        self.write(f'TRIG1:SOUR {channel}')
        self.write('TRIG1:TYPE WIDT')
        self.write(f'TRIG1:WIDT:POL {width_polarity}') # POS | NEG
        self.write(f'TRIG1:WIDT:RANG {width_range}') # WITHin | OUTSide | SHORter | LONGer
        self.write(f'TRIG1:WIDT:WIDT {width}')  ## Range: 100E-12 to 10000
                                                # Increment: 100E-9
                                                # *RST: 5E-9
                                                # Default unit: s
        self.write(f'TRIG1:WIDT:DELT {delta}')  ## Range: 0 to 432
                                                # Increment: 500E-12
                                                # *RST: 0
                                                # Default unit: s

    def timeout_trigger(self, trigger_channel, timeout_range='HIGH', timeout_time=1E-3):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4
        self.write(f'TRIG1:SOUR {channel}')
        self.write(f'TRIG1:TYPE TIM')
        self.write(f'TRIG1:TIM:RANG {timeout_range}') # HIGH | LOW | EITHer
        self.write(f'TRIG1:TIM:TIME {timeout_time}') # Range: 100E-12 to 10000
                                                        # Increment: 100E-9
                                                        # *RST: 100E-9
                                                        # Default unit: s

    def trigger_level(self, trigger_channel, trigger_level):
        self.write(f'TRIG1:LEV{trigger_channel} {trigger_level}') # range: -10 to 10, increment: 1E-3

    def trigger_mode(self, mode):
        self.write(f'TRIG:MODE {mode}') # AUTO | NORMal | FREerun

    def force_trigger(self):
        self.write('TRIG:FORC')

    def trigger_status(self):
        status = self.write('ACQ:CURR?')
        status = int(status)
        return status

    ### Channel Settings ###

    def channel_BW(self, channel, channel_BW):
        """
        500: 'FULL'
        20: 'B20'
        200: 'B200'
        """
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        if channel_BW == 500:
            channel_BW = 'FULL'
        elif channel_BW == 20:
            channel_BW = 'B20'
        elif channel_BW == 200:
            channel_BW = 'B200'

        self.write(f'CHAN{channel}:BAND {channel_BW}')

    def channel_offset(self, channel, channel_offset):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:OFFS {channel_offset}')

    def channel_position(self, channel, channel_position):
        self.channel_offset(channel, 0)

        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:POS {channel_position}')

    def channel_coupling(self, channel, channel_coupling):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:COUP {channel_coupling}') # DC | DCLimit | AC

    def channel_scale(self, channel, channel_scale):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:SCAL {channel_scale}') # V/div

    def channel_state(self, channel, state='OFF'):
        self.write(f"CHANnel{channel}:STATe {state}")
        self.write(f"MEASurement{channel}:ENABle {state}")


    def channel_settings(self, state, channel=1, scale=1, position=0, label='IOUT', color='LIGHT_BLUE', rel_x_position=50, bandwidth=500, coupling='DCLimit', offset=0):
        """
        COLOR OPTIONS:

        - LIGHT_BLUE
        - YELLOW
        - PINK
        - GREEN
        - BLUE
        - ORANGE


        returns : state of the channel ('ON' or 'OFF')
        """

        self.channel_state(channel, state)

        self.channel_position(channel, position)
        self.channel_scale(channel, scale)
        
        if state == 'ON': print(f"CH{channel} - {label}")
        self.channel_label(channel, label, rel_x_position)
        self.channel_color(channel, color)

        self.channel_BW(channel, bandwidth)
        
        self.channel_coupling(channel, coupling)
        self.channel_offset(channel, offset)
        
        self.display_intensity()

        self.measure_enable(channel, state)

        return state

    def channel_label(self, channel, label, rel_x_position):
        self.write(f"DISPlay:SIGNal:LABel:REMove 'Label1', C{channel}W1")
        self.write(f"DISPlay:SIGNal:LABel:ADD 'Label1', C{channel}W1, '{label}', REL, {rel_x_position}, 0.1")
    
    def query_channel_label(self, channel):
        return self.write(f"DISPlay:SIGNal:LABel:TEXT? 'Label1', C{channel}W1")
    
    def channel_color(self, channel, color):

        light_blue = int('ff26e9ff', 16)
        yellow = int('ffffff00', 16)
        pink = int('ffff49fb', 16)
        green = int('ff16e500', 16)
        blue = int('ff0080ff', 16)
        orange = int('ffff6000', 16)

        colors = {  'LIGHT_BLUE':light_blue,
                    'YELLOW': yellow,
                    'PINK': pink,
                    'GREEN': green,
                    'BLUE': blue,
                    'ORANGE': orange
        }

        signal = {
            '1': 2,
            '2': 5,
            '3': 8,
            '4': 11
        }

        self.write(f"DISPlay:COLor:SIGNal{signal[str(channel)]}:COLor {colors[color]}")
    
    def display_intensity(self, intensity=100):
        self.write(f"DISPlay:INTensity 0")
        self.write(f"DISPlay:INTensity {intensity}")

    def query_cursor_state(self, cursor_set):
        return int(self.write(f"CURS{cursor_set}:STAT?"))
    
    def query_cursor_source(self, cursor_set):
        return self.write(f"CURS{cursor_set}:SOUR?")

    def cursor(self, channel=1, cursor_set=1, X1=1, X2=1, Y1=0, Y2=0, type='VERT'):
        self.write(f"CURSor{cursor_set}:FUNCtion {type}") # HORizontal | VERTical | PAIRed
        self.write(f"CURS{cursor_set}:STAT ON")
        self.write(f"CURSor{cursor_set}:TRACking OFF")
        self.write(f"CURS{cursor_set}:SOUR C{channel}W1")
        self.write(f"CURS{cursor_set}:X1P {X1}")
        self.write(f"CURS{cursor_set}:X2P {X2}")
        self.write(f"CURS{cursor_set}:Y1P {Y1}")
        self.write(f"CURS{cursor_set}:Y2P {Y2}")
        

    """MEASURE"""

    def measure_enable(self, channel, state='ON'):
        self.write(f"MEASurement{channel}:ENABle {state}")

    def measure_source(self, channel):
        self.write(f"MEASurement{channel}:SOURce C{channel}W1")
        self.write(f"MEASurement{channel}:CATegory AMPTime")
    
    def measure_off(self, channel):
        self.write(f"MEASurement{channel}:AOFF")

    def measure(self, channel, measure_list):
        """
        HIGH | LOW | AMPLitude | MAXimum | MINimum | PDELta |
        MEAN | RMS | STDDev | POVershoot | NOVershoot | AREA |
        RTIMe | FTIMe | PPULse | NPULse | PERiod | FREQuency |
        PDCYcle | NDCYcle | CYCarea | CYCMean | CYCRms |
        CYCStddev | PULCnt | DELay | PHASe | BWIDth | PSWitching |
        NSWitching | PULSetrain | EDGecount | SHT | SHR | DTOTrigger |
        PROBemeter | SLERising | SLEFalling
        """

        # self.measure_enable(channel, state='ON')
        self.measure_source(channel)
        self.measure_off(channel)
        measure_list = measure_list.strip(" ").split(",")
        for type in measure_list:
            self.write(f"MEASurement{channel}:ADDitional {type}, ON")

    def active_measure(self, channel):
        # self.write(f":MEASurement:ARNames {channel}")
        print(self.write(f":MEASurement:ARES? {channel}"))
        
    def cleanup(self):
        self.close()





"""Charles' Custom Functions"""

# initialize variables
global Iout_index
global waveform_counter
global start
global waveforms_folder
waveform_counter = 0
Iout_index = 0
from datetime import datetime



import os
import shutil

def path_maker(file_path: str):
    """
        file_path: Enter the file path you want to create

        returns the string value of new path created
    """

    folder_list = file_path.split('/')

    new_path = ' '
    for i in folder_list:
        
        if new_path == ' ':
            path = f'{i}/'
        else: 
            path = new_path + f'{i}/'

        if not os.path.exists(path):
            os.mkdir(path)
            
        new_path = path
    return new_path


def remove_file(file_path: str):
    """
        file_path: Enter the file path you want to delete

        returns the string value of new path created
    """

    if os.path.exists(file_path): os.remove(file_path)


def move_file(source_path: str, destination_path: str):
    """
        source_path : path/file
        destination_path : path/file
    """
    source = f'{source_path}'
    destination = f'{destination_path}'
    remove_file(destination)
    shutil.move(source, destination)


def create_sfx(message):
    language = 'en'
    myobj = gTTS(text=message, lang=language, slow=False)
    path_maker(f'{os.getcwd()}/sfx')
    if not os.path.exists(f"{os.getcwd()}/sfx/{message}.mp3"):
        myobj.save(f"{message}.mp3")
        move_file(f"{message}.mp3", f"{os.getcwd()}/sfx/{message}.mp3")
    playsound(f"{os.getcwd()}/sfx/{message}.mp3")

def tts(message):
    """
        function: translates text to speech based on user's input (message)
    """
    create_sfx(message)
    print(f"{message}")

def prompt(message):
    """
        function: translates text to speech based on user's input (message) and wait for user intervention to continue program
    """
    create_sfx(message)
    input(f"{message}")

def start_timer():
    start = datetime.now()
    return start

def end_timer(start):
    end = datetime.now()
    total_time = end-start
    temp = ""
    
    print(f"START TIME: {temp:>12}{start}")
    print(f"END TIME: {temp:>12}{end}")
    print(f"TOTAL TIME: {temp:>12}{total_time}")


def headers(test_name):
    global start
    print()
    print("="*80)
    print(f"Test: {test_name}")
    # create_folder(test_name)
    
    start = datetime.now()
    print("="*80)
    # tts(f"Starting {test_name} test.")

def create_folder(test_name):
    if not os.path.exists('waveforms'): os.mkdir('waveforms')
    waveforms_folder = f'waveforms/{test_name}'
    pathname = f"{os.getcwd()}/{waveforms_folder}"
    if not os.path.exists(pathname): os.mkdir(pathname)

def sfx():
    import winsound as ws
    ws.PlaySound("dingding.wav", ws.SND_ASYNC)
    sleep(2)  

def footers(waveform_counter):

    print("="*80)
    if waveform_counter > 0: print(f'{waveform_counter} waveforms captured.')
    print('test complete.')
    print()
    end = datetime.now()
    total_time = end-start
    temp = ""
    
    print(f"START TIME: {temp:>12}{start}")
    print(f"COMPLETION TIME: {temp:>7}{end}")
    print(f"TOTAL TESTING TIME: {temp:>16}{total_time}")
    print("="*80) 
    
def soak(soak_time):
    for seconds in range(soak_time, 0, -1):
        sleep(1)
        print(f"{seconds:5d}s", end="\r")
    print("       ", end="\r")

def convert_argv_to_int_list(a=[]):
    str_list = a.strip('[]').split(',')
    int_list = [int(x) for x in str_list]
    return int_list

def convert_argv_to_str_list(a=[]):
    str_list = a.strip('[]').split(',')
    return str_list

def roundup(x):
    import math
    return int(math.ceil(x / 100.0)) * 100


import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt

def get_anchor(col, row):
    """
    get anchor given a numerical col row  -> (col = 2, row = 4 -> 'B4')
    returns anchor (str)
    """
    anchor = f"{get_column_letter(col)}{row}"
    return anchor

def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)
    # return pd.read_csv(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """



    sheet_list = wb.get_sheet_names()
    if sheet_name not in sheet_list: wb.create_sheet(sheet_name)
    try:
        default_sheet = wb.get_sheet_by_name('Sheet1')
        wb.remove_sheet(default_sheet)
    except: pass

    # print(sheet_list)
    if "Chart" not in sheet_list: wb.create_sheet("Chart")
    

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1

    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]
            try: wb[sheet_name][f'{get_column_letter(col)}{row}'].style = highlight
            except: wb[sheet_name][f'{get_column_letter(col)}{row}'].style = 'highlight'

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39 rows, 16 columns
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = folder_path + filename
    # file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1056
    img.height = 659.90551181

    sheet_list = wb.get_sheet_names()
    if sheet_name not in sheet_list: wb.create_sheet(sheet_name)
    try:
        default_sheet = wb.get_sheet_by_name('Sheet1')
        wb.remove_sheet(default_sheet)
    except: pass

    

    

    col, row = col_row_extractor(anchor)
    wb[sheet_name][f'{get_column_letter(col)}{row-1}'] = filename
    wb[sheet_name].add_image(img)


#################### DATA EXPORT CODES ###########################################################
def create_header_list(df_header_list):
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    # print(df)
    return df

def export_to_excel(df, waveforms_folder, output_list, excel_name, sheet_name, anchor):
    # import datetime
    # now = datetime.datetime.now()
    # now.strftime('%m-%d-%y %H:%M:%S')
    # print(now)
    df.loc[len(df)] = output_list
    # print(df.loc[1:1E5])

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)

def export_df_to_excel(df, waveforms_folder, excel_name, sheet_name, anchor):

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)

def export_screenshot_to_excel(excel_name, waveforms_folder, sheet_name, filename, anchor):

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    image_to_excel(wb, sheet_name, filename=filename, folder_path=waveforms_folder, anchor=anchor)
    wb.save(dst)


























from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

def create_scatter_chart(title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

    chart = ScatterChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale

    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.minorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.y_axis.minorUnit = y_minor_unit


    return chart

def create_bar_chart(title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

    chart = BarChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15
    chart.type = "col"

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale

    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.minorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.y_axis.minorUnit = y_minor_unit


    # data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
    # cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    # chart1.add_data(data, titles_from_data=True)
    # chart1.set_categories(cats)
    # chart1.shape = 4
    # ws.add_chart(chart1, "A10")


    return chart

def reset_chartsheet(wb):
    chart_sheet = wb["Chart"]
    try:
        no_of_existing_charts = len(chart_sheet._charts)
        for i in range(no_of_existing_charts):
            del chart_sheet._charts[(i-1)]
    except: pass
    
    return chart_sheet

def save_chartsheet(chart_sheet, chart, chart_position):
    chart_sheet.add_chart(chart, chart_position)

def append_series(path, wb, ws_name, x_anchor, last_row_anchor, series_title, chart):
    ## used in der-727_chart_compiler.py
    
    df = excel_to_df(path, ws_name, x_anchor, last_row_anchor)
    ws = wb[ws_name]
    xcol, xrow = col_row_extractor(x_anchor)
    last_col, last_row = col_row_extractor(last_row_anchor)
    xvalues = Reference(ws, min_col=xcol, min_row=xrow, max_row=last_row)
    values = Reference(ws, min_col=last_col, min_row=xrow, max_row=last_row)
    series = Series(values, xvalues, title=series_title)
    series.marker=openpyxl.chart.marker.Marker('auto')
    series.graphicalProperties.line.noFill=False
    chart.series.append(series) 
















if __name__ == '__main__':
    pass


