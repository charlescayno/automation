from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

# File paths
source_file = "file.xlsx"
dest_file = "destination.xlsx"
source_file_path = f"C:/Users/ccayno/Documents/Charles/Work/DER/DER-1081/07 - Test Data/Rev B_E1_own_made_5.5lu/Efficiency vs Input Voltage/11-25-2025/15V/{source_file}"
dest_file_path = f"C:/Users/ccayno/Documents/Charles/Work/DER/DER-1081/07 - Test Data/Rev B_E1_own_made_5.5lu/Efficiency vs Input Voltage/11-25-2025/15V/{dest_file}"

# Load workbooks
src_wb = load_workbook(source_file_path)
dst_wb = load_workbook(dest_file_path)

# Select sheets
src_sheet = src_wb["Efficiency vs Input Voltage"]      # change sheet name if needed
dst_sheet = dst_wb["Sheet 1"]      # change sheet name if needed

# Function to convert Excel address to row/col indices
def get_range_indices(start_cell, end_cell):
    start_col_letter, start_row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(start_col_letter)

    end_col_letter, end_row = coordinate_from_string(end_cell)
    end_col = column_index_from_string(end_col_letter)

    return start_row, end_row, start_col, end_col

# Source range
source_start = "A2"
source_end   = "T11"
src_start_row, src_end_row, src_start_col, src_end_col = get_range_indices(source_start, source_end)

# Calculate source size
num_rows = src_end_row - src_start_row + 1   # number of rows in source range
num_cols = src_end_col - src_start_col + 1   # number of columns in source range

# Destination start anchor
dest_start = "A68"
dest_start_row, dest_start_col = get_range_indices(dest_start, dest_start)[0:2]

# Calculate destination end based on source size
dest_end_row = dest_start_row + num_rows - 1
dest_end_col = dest_start_col + num_cols - 1
dest_end_cell = f"{get_column_letter(dest_end_col)}{dest_end_row}"

print(f"Destination range: {dest_start}:{dest_end_cell}")  # Example: A68:T77
input()

# # Copying values
# for i, row in enumerate(range(src_start_row, src_end_row + 1)):
#     for j, col in enumerate(range(src_start_col, src_end_col + 1)):
#         src_cell = src_sheet.cell(row=row, column=col)
#         dst_cell = dst_sheet.cell(row=dest_start_row + i, column=dest_start_col + j)
#         dst_cell.value = src_cell.value


# Copying values from source to destination

for i, row in enumerate(range(src_start_row, src_end_row + 1)):
    for j, col in enumerate(range(src_start_col, src_end_col + 1)):
        src_cell = src_sheet.cell(row=row, column=col)
        dst_cell = dst_sheet.cell(row=dest_start_row + i, column=dest_start_col + j)
        dst_cell.value = src_cell.value  # Copy value only


# Save the destination file
dst_wb.save(dest_file_path)

print("Copied successfully!")
