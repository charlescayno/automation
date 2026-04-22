from misc_codes.general_settings import *
########################################## USER INPUT ##########################################
input_path = r"C:\Users\ccayno\Desktop\DER-999/cmc/120Vac/"


output_filename = "summary.xlsx"

output_path = input_path + '/' + output_filename

# test = f"Chart Compiler"
# waveforms_folder = GENERAL_FUNCTIONS().CREATE_PATH(project, test)
########################################## USER INPUT ##########################################

def main():

    # excel_files_list = GENERAL_FUNCTIONS().EXCEL_FILES_LIST(input_path)
    # print(excel_files_list)

    filename = f"Unit 0, 120Vac.xlsx"

    import pandas as pd
    from openpyxl import load_workbook

    # Load the Excel file
    file_path = input_path + f'/{filename}'
    workbook = load_workbook(filename=file_path)

    # Get the sheet names
    sheet_names = workbook.sheetnames

    # Choose the sheet to read from
    sheet = workbook[sheet_names[0]]  # Assuming you want to read from the first sheet

    # Read data into a DataFrame
    data = sheet.values
    columns = next(data)  # Assumes the first row contains column names
    df = pd.DataFrame(data, columns=columns)

    # Remove NaN or blank values
    df = df.dropna()  # Drop rows with any NaN value
    df = df.dropna(how='all')  # Drop rows with all NaN values
    df = df.dropna(subset=columns)  # Drop rows with NaN values in specific columns

    # Reset the index if necessary
    df = df.reset_index(drop=True)

    # Print the DataFrame
    print(df)

    import pandas as pd
    import matplotlib.pyplot as plt

    # Assuming you already have a DataFrame called 'df'

    # Line plot
    # plt.plot(df['Timestamp'], df['Eff (%)'])
    # plt.plot(df['Timestamp'], df['Vo (V)'])
    plt.plot(df['Timestamp'], df['Io1 (A)'])
    plt.xlabel('X-axis')
    plt.ylabel('Y-axis')
    plt.title('Line Plot')
    plt.show()

    input()

    

if __name__ == "__main__":
    # headers(test)
    main()
    # footers(waveform_counter)
