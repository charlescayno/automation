import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

# Set the input path and file name
input_path = r"C:/Users/ccayno/Desktop/DER-999/FOD Characterization new"
file_name = "DER-999 120 Vac - Unit 1.xlsx"

# Load the workbook and select the active sheet
wb = load_workbook(filename=f'{input_path}/{file_name}')
ws = wb.active

# Create a list of column headers for the data frame, excluding empty cells
headers = []
for cell in ws[2]:
    if cell.value:
        headers.append(cell.value)

# Read data from the active sheet into a data frame, excluding empty columns
df = pd.read_excel(f'{input_path}/{file_name}', header=1, usecols=headers[1:]).dropna(how='all')
print(df)

x_col = 'CHS'
x = df[x_col]
print(x)

# Create a list of y columns
y_columns = [col for col in df.columns if col != x_col and col != "Unnamed: 0" and col != "CHS"]

# Delete the existing "Chart Summary" sheet
if "Chart Summary" in wb.sheetnames:
    del wb["Chart Summary"]

# Create a new "Chart Summary" sheet and set up the charts
summary_ws = wb.create_sheet("Chart Summary")

for i, y_col in enumerate(y_columns):
    # Create a new scatter chart
    chart = ScatterChart()
    chart.title = y_col
    chart.x_axis.title = x_col

    # Add the x and y data to the chart
    xdata = Reference(ws, min_col=df.columns.get_loc(x_col)+3, min_row=2, max_row=len(df)+2)
    ydata = Reference(ws, min_col=df.columns.get_loc(y_col)+3, min_row=2, max_row=len(df)+2)
    series = Series(ydata, xdata)
    chart.series.append(series)

    # Set the axis labels
    chart.y_axis.title = y_col

    # Add the chart to the "Chart Summary" sheet
    summary_ws.add_chart(chart, f'B{(i+1)*15+2}')

# Save the changes and print a notification
wb.save(f'{input_path}/{file_name}')
print("Charts have been created!")
