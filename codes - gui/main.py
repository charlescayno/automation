import os
import sys
from unittest import result

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox
from QtDesigner import qmlview

from ui.main_window import Ui_Form
from qt_material import apply_stylesheet

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import(
    ScatterChart,
    Reference,
    Series
)

from pandas import read_csv

# https://www.youtube.com/watch?v=XXPNpdaK9WA&t=645s
class QuadramaxCC(QtWidgets.QWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.ui.pushButton_cancel.clicked.connect(self.close)
        self.ui.pushButton_browse.clicked.connect(self.open_folder_dialog)
        self.ui.pushButton_ok.clicked.connect(self.on_okPushButton_clicked)

        self.msg_box = QMessageBox()

        self.source_directory = ''
        self.cc_plot_filenames_list = []

        self.wb = Workbook()


    def open_folder_dialog(self):
        username = os.getlogin()
        default_path = f"C:\\Users\\{username}\\Desktop"
        dialog = QtWidgets.QFileDialog()

        # C:\Users\rnueda\Desktop\Delete\Quadramax 100W\SourceTests_220506_222642\SourceTests_220506_222642
        source_directory = dialog.getExistingDirectory(self, 'Select Quadramax Results Directory', default_path)

        if not source_directory == '':
            self.ui.lineEdit_folder_path.setText(source_directory)
            self.source_directory = source_directory

    def on_okPushButton_clicked(self):

        # Check if there are iPpsTolerancePoints csv files
        #
        valid_directory = self.check_if_valid_directory()

        if valid_directory:
            # copy csv files to sheets
            self.copy_csv_files_to_excel()

            # generate plots
            self.generate_scatter_plots()
            
            # save the file
            file_saved = False
            output_filepath = os.path.join(self.source_directory, "CC Plots.xlsx")
            try:
                del self.wb['Sheet']
                self.wb.save(output_filepath)
                file_saved = True
            except PermissionError:
                self.msg_box_info(title="Permission Error", message="Please close the CC Plots file for the selected folder.")

            # open the file
            if file_saved:
                os.system("\"" + output_filepath + "\"")
            
            # exit the program
            raise SystemExit


    def check_if_valid_directory(self):
        if os.path.exists(self.source_directory):
            # Loop through file names, search for iPpsTolerancePoints

            dir_items = os.listdir(self.source_directory)

            # search directory items if there are iPpsTolerancePoints .csv files
            for filename in dir_items:
                
                if filename[:20] == "iPpsTolerancePoints_" and filename [-4:]=='.csv':
                    # filepath = os.path.join(self.source_directory, filename)
                    self.cc_plot_filenames_list.append(filename)

            if len(self.cc_plot_filenames_list) == 0:
                self.msg_box_info(title="Directory Error", message="Selected folder does not contain the iPpsTolerancePoints.csv files")
                return False
            else:
                return True
        else:
            self.msg_box_info(title="Directory Error", message="Please enter a valid directory")
            return False

    def msg_box_info(self, title, message):
            self.msg_box.setIcon(QMessageBox.Information)
            self.msg_box.setText(message)
            self.msg_box.setWindowTitle(title)
            self.msg_box.setStandardButtons(QMessageBox.Ok)
            self.msg_box.exec_()

    
    
    def copy_csv_files_to_excel(self):
        filenames = self.cc_plot_filenames_list

        for filename in filenames:
            self.wb.create_sheet(filename)
            data = read_csv(os.path.join(self.source_directory, filename),
                names=["test_num","pass/fail","Vout","Iout Min Lim","Iout Meas","Iout Max Lim"])

            del data["test_num"]
            del data["pass/fail"]

            ws = self.wb[filename]  

            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)


    def generate_scatter_plots(self):
        for sheet in self.wb.worksheets:
            chart = ScatterChart()
            chart.title = sheet.title
            chart.x_axis.title = "Output Current (A)"
            chart.y_axis.title = "Output Voltage (V)"

        
            values = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            for i in range(2,5):
                xvalues = Reference(sheet, min_col=i, min_row=2, max_row=sheet.max_row)
                series = Series(values, xvalues, title=sheet.cell(row=1, column=i).value)
                # series.title = ws.cell(row=1, column=i).value
                chart.series.append(series)

            sheet.add_chart(chart, "F1")
        
        self.msg_box_info(title="Charts", message="Charts have been generated.")
        self.ui.lineEdit_folder_path.setText("")
            
if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    
    widget = QuadramaxCC()
    apply_stylesheet(app, theme='dark_teal.xml')
    widget.show()

    app.exec_()